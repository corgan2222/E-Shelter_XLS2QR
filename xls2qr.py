#!/usr/bin/env python3

import os
import argparse
import logging
import sys
import export
import util
import logging
from rich.logging import RichHandler
from PyPDF2 import PdfFileMerger, PdfFileReader

import openpyxl
import createQR

LOG_LEVELS = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]
DEFAULT_LOG_LEVEL = "INFO"

    
def runExcelReader(xlsFile, output_path, svg_path, pdf_path, sheet="Vossloh und Schwabe Multisensor"):  

    if not os.path.exists(xlsFile):
        log.error("[bold red blink]Excel File not found " + xlsFile + " [/]", extra={"markup": True})          
        log.error("Exit")  


    wb_obj = openpyxl.load_workbook(xlsFile)
    sheet_obj = wb_obj[sheet]
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    for i in range(2, m_row + 1):
        eshellid_obj = sheet_obj.cell(row = i, column = 1)
        label_obj = sheet_obj.cell(row = i, column = 2)
        gw_obj = sheet_obj.cell(row = i, column = 3)
        floor_obj = sheet_obj.cell(row = i, column = 4)
        room_obj = sheet_obj.cell(row = i, column = 5)
        id_obj = sheet_obj.cell(row = i, column = 6)
        qr_obj = sheet_obj.cell(row = i, column = 7)        

        try:
            if len(gw_obj.value) > 0: 
                svg_path2 = export.createFolderPath(svg_path, gw_obj.value)
                pdf_path2 = export.createFolderPath(pdf_path, gw_obj.value)
                svg_filename_label = os.path.join(svg_path2, label_obj.value + "_" + id_obj.value + ".svg")
                pdf_filename_label = os.path.join(pdf_path2, label_obj.value + "_" + id_obj.value + ".pdf")
                log.debug("svg_filename: " + svg_filename_label)
                log.debug("pdf_filename: " + pdf_filename_label)
        except:
            pass    

        try:
            if len(gw_obj.value) > 0:
                log.info("eshellid: " + str(eshellid_obj.value) + " | label: " + str(label_obj.value) + " | gw: " + str(gw_obj.value) + 
                    " | floor: " + str(floor_obj.value) + " | room: " + str(room_obj.value) + " | id: " + str(id_obj.value) + 
                    " | qr: " + str(qr_obj.value)  )   
            createQR.make_pdf_file(svg_filename_label, pdf_filename_label, str(qr_obj.value), str(floor_obj.value), str(room_obj.value), str(label_obj.value), str(gw_obj.value), str(eshellid_obj.value), str(id_obj.value))  
        except:
            pass    

    #merge all PDF to one PDF per Gateway        
    mergePDFs(pdf_path,"pdf",output_path)
    log.info("[bold green blink]- Finished! " + str(xlsFile) + " [/]", extra={"markup": True})



def mergePDFs(pdf_path,fileextension, output_path):
    pdf_folders = util.fast_scandir(pdf_path)
    for folder in pdf_folders:
        log.debug("Scan Folder for PDF: " + folder )

        # Call the PdfFileMerger
        mergedObject = PdfFileMerger()

        for filename in os.listdir(folder):
            if filename.endswith(fileextension) :        
                pdf = os.path.join(folder, filename)
                log.debug("Read ODF File: " + pdf )
                mergedObject.append(PdfFileReader(pdf, 'rb'))

        exportPDF_Foldername = os.path.basename(folder)
        exportPDF = os.path.join(output_path, exportPDF_Foldername + ".pdf") 
        log.info("Export PDF to: " + exportPDF)
        mergedObject.write(exportPDF)


def allFilesMode(path,fileextension,output_path,svg_path,pdf_path):
    #Reads all files from Folder          
    for filename in os.listdir(path):
        if filename.endswith(fileextension) :        
            log.info("Process File: " + os.path.join(path, filename))
            runExcelReader(os.path.join(path,filename), output_path, svg_path, pdf_path, "Vossloh und Schwabe Multisensor")
        else:
            log.info("Finished. No more Files found in:" + path)
            continue  

def get_parser():
    """Get parser object """
    from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter
    parser = argparse.ArgumentParser(description='Understand functioning')
    parser.add_argument("-x", "--xlsx", type=str, required=False,  metavar="FILE", help="Path to Excel File", default="") 
    parser.add_argument("-o", "--outputfolder", type=str, required=False,  metavar="FILE", help="Path to output folder", default="") 
    parser.add_argument("-s", "--xlsfolder", type=str, required=False,  metavar="FILE", help="Path to Excel Files Folder", default="") 
    parser.add_argument("-l", "--logfile", type=str, required=False, help="Logfile", default="")
    parser.add_argument("--verbose", "-v", dest="log_level", action="append_const", const=-1,)
    parser.add_argument("--quiet", "-q", dest="log_level", action="append_const", const=1,)
    return parser

def setLogging(args):   
    log_level = LOG_LEVELS.index(DEFAULT_LOG_LEVEL)
    for adjustment in args.log_level or ():
            log_level = min(len(LOG_LEVELS) - 1, max(log_level + adjustment, 0))

    log_level_name = LOG_LEVELS[log_level]

    global log
    log = logging.getLogger(__name__)    
    shell_handler = RichHandler()    
    log.setLevel(log_level_name)
    shell_handler.setLevel(log_level_name)    
    fmt_shell = '%(message)s'    
    shell_formatter = logging.Formatter(fmt_shell)
    shell_handler.setFormatter(shell_formatter)
    log.addHandler(shell_handler)
    
    if len(args.logfile) > 0:
        file_handler = logging.FileHandler(args.logfile)
        file_handler.setLevel(logging.DEBUG)
        fmt_file = '%(levelname)s %(asctime)s [%(filename)s:%(funcName)s:%(lineno)d] %(message)s'
        file_formatter = logging.Formatter(fmt_file)
        file_handler.setFormatter(file_formatter)
        log.addHandler(file_handler)        

def main(args):

    setLogging(args)
    log.info("[bold green]ExcelData -> QR Code  © Stefan Knaak - E-Shelter Security.  [/]", extra={"markup": True}) 
    
    #Folders
    cwd = os.getcwd()
    xls_path = export.checkInputFolderPath(args.xlsfolder, "xlsx")
    output_path = export.checkInputFolderPath(args.outputfolder, "output")
    svg_path = export.checkInputFolderPath(args.outputfolder, "output_svg")
    pdf_path = export.checkInputFolderPath(args.outputfolder, "output_pdf")

    log.debug("xls_path: " + str(xls_path))    
    log.debug("output: " + str(output_path))    
    log.debug("svg_path: " + str(svg_path))    
    log.debug("pdf_path: " + str(pdf_path))    
    
    #Operation Mode
    #Get Excel File from Command line
    if len(args.xlsx) > 0 :
        if os.path.isfile(args.xlsx) :
            log.info("[bold purple]Single File Mode.  [/]", extra={"markup": True})  
            log.info("[bold green]Image: " + args.xlsx + " [/]", extra={"markup": True})  
            runExcelReader(args.xlsx, output_path, svg_path, pdf_path, "Vossloh und Schwabe Multisensor")
        else:  
            log.info("[bold cyan]Single File Mode.  [/]", extra={"markup": True})  
            if not os.path.isfile(args.xlsx):
                log.error("[bold red]Excel File not found [/] [yellow] " + args.xlsx + " [/]", extra={"markup": True})  
            log.error("Exit")    

    else: #/data/xlsx folder
        log.info("[bold cyan]Folder Mode[/]", extra={"markup": True}) 
        if len(args.xlsfolder) > 0:
            log.info("[bold white]Folder: [/] " + xls_path , extra={"markup": True}) 

        if not os.path.exists(xls_path):
            log.error("[bold red blink]xlsx Folder not found " + xls_path + " [/]", extra={"markup": True})  
            log.error("Exit")  
        else:  
            log.info("[bold cyan]Process all Excel Files from " + xls_path + " [/]", extra={"markup": True})     
            allFilesMode(xls_path,'xlsx',output_path,svg_path,pdf_path)       
              
if __name__ == "__main__":    
    args = get_parser().parse_args()
    try:
        main(args)
    except BaseException:
        import sys
        print(sys.exc_info()[0])
        import traceback
        print(traceback.format_exc())
    finally:
        print("Press Enter to continue ...")
        input()
